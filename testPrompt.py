import json
import pandas as pd
from openpyxl import load_workbook

def corpusToScript(code):
    post_fix = code[4:]
    post_fix_num = int(post_fix)
    return post_fix_num

def getSeries(id):
    prefix = "poly"
    id_str = str(id)
    id_len = len(id_str)
    series = prefix + '0'*(5-id_len) + id_str
    return str(series)

def concatData(df_current, position_count, pos):
    for index1, row1 in df_current.iterrows():
        #for each attribute (dict), assign the current store_id, its current position, and id & name
        df_current._set_value(index1, "Script Num", row["Script Num"])
        df_current._set_value(index1, "Script Name", bus_prefix + str(row["Script Num"]))
        curr_dict = row1["Attributes"]
        curr_dict["store_id"] = curr_poly["store_id"][pos]
        curr_dict["Position"] = position_count
        df_current._set_value(index1, "Attributes", str(curr_dict))
        position_count += 1

if __name__ == "__main__":
    #"global" variables
    dataTable = "D:\Jun22\output\Test_prompt.xlsx"
    inputSampleTable = "D:\Jun22\output\prompt_test_base_input.xlsx"
    bus_prefix = "BusinessSurvey"

    # file with json (polygon & store info table)
    wb_org = load_workbook(dataTable)
    ws_org = wb_org.active
    data_org = ws_org.values
    column_org = next(data_org)[0:]
    df_org = pd.DataFrame(data_org, columns=column_org)
    df_org = df_org.rename(columns={"Corpus Code": "Script Num"})
    #change column name and data type for table joining
    for index, row in df_org.iterrows():
        poly_id_org = row["Script Num"]
        poly_id = corpusToScript(poly_id_org)
        df_org._set_value(index, "Script Num", poly_id)

    # input table file
    wb_input = load_workbook(inputSampleTable)
    ws_input = wb_input.active
    data_input = ws_input.values
    column_input = next(data_input)[0:]
    df_input = pd.DataFrame(data_input, columns=column_input)
    #change data type of "Attributes" from string to json object (dictionary)
    for index, row in df_input.iterrows():
        converted = json.loads(row["Attributes"])
        df_input._set_value(index, "Attributes", converted)
    pd.set_option('max_columns', None)

    #for input table, alter the "Script Num", "Script Name", "Attributes" columns from the polygon & store info table
    sample_size = df_input.shape[0]
    df_final = pd.DataFrame(columns=df_input.columns)
    for index, row in df_org.iterrows():
        position_count = 1
        #make a deep copy of the input sample table
        df_current = df_input.copy(deep=True)
        curr_poly = json.loads(row["Script Attributes"])
        store_num = len(curr_poly["store_id"])
        #if there's only 1 store in the polygon
        if store_num == 1:
            concatData(df_current, position_count, 0)
            df_final = pd.concat([df_final, df_current]) #concat the sample table (after assigning values) to the big table
            df_current = df_current.iloc[0:0] #clear the sample table
        # if there're only multiple stores in the polygon
        else:
            #for each store, repeat the operation above
            for i in range(store_num):
                df_current = df_input.copy(deep=True)
                concatData(df_current, position_count, i)
                df_final = pd.concat([df_final, df_current])
                df_current = df_current.iloc[0:0]

    #generate the automatically increasing id (corpus code)
    df_final = df_final.reset_index(drop=True)
    corpus_count = 1
    for index, row in df_final.iterrows():
        df_final._set_value(index, "Corpus Code", getSeries(corpus_count))
        corpus_count += 1

    #set some columns with default values
    df_final["Repeat Times"] = 5
    df_final["Duration"] = 43200
    df_final["Project ID"] = 223
    df_final["Project Num"] = 741

    #merge the two tables and output to excel
    df_all = pd.merge(df_final, df_org, on="Script Num", how="inner")
    df_all.to_excel("D:\Jun22\output\prompt_test_base_output.xlsx", sheet_name="sheet1", index=False)

    #Note: For the Attributes column in the output excel, do the following operations:
    #      (1) ' to "
    #      (2) Can"t Tell to Can't Tell
    #      (3) True to true
