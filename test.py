import json
from openpyxl import load_workbook, Workbook
import pandas as pd

#Depend on the length of the id, add 0's to the polygon_id
def changeKeyName(my_key):
    prefix = my_key[0:4]
    real_num = my_key[8:]
    ret = ""
    if len(real_num) == 4:
        ret = prefix + '0' + real_num
    if len(real_num) == 3:
        ret = prefix + "00" + real_num
    return ret

if __name__ == '__main__':
    #assign file path, then read file
    leftTable = "D:\Jun22\output\All Jakarta Remaining_polygons_left.xlsx" #store info
    rightTable = "D:\Jun22\output\All Jakarta Remaining_polygons_right.xlsx" #polygon info (coordinates)
    wb_left = load_workbook(leftTable)
    ws_left = wb_left.active
    wb_right = load_workbook(rightTable)
    ws_right = wb_right.active

    #select targeted columns
    data_left = ws_left.values
    column_left = next(data_left)[0:]
    df_left = pd.DataFrame(data_left, columns=column_left)
    df_left = df_left.iloc[:, [5,6,7,8,9,10,12]]

    #data cleaning of the coordinate table
    count_row = 1
    for row in ws_right:
        if count_row >= 2:
            count_col = 1
            x_sum = 0
            y_sum = 0
            #remove left bracket in x coordinate cell, right bracket in y coordinate cell, change type to float
            for cell in row:
                if not(count_col == 1 or count_col == 10):
                    if count_col % 2 == 0:
                        real_val = cell.internal_value[1:]
                        x_sum += float(real_val)
                        ws_right.cell(row=count_row, column=count_col).value = float(real_val)
                    else:
                        real_val = cell.internal_value[1:-1]
                        y_sum += float(real_val)
                        ws_right.cell(row=count_row, column=count_col).value = float(real_val)
                count_col += 1
            #calculate average latitute and longitude
            ws_right.cell(row=count_row, column=11).value = x_sum / 4
            ws_right.cell(row=count_row, column=12).value = y_sum / 4
        count_row += 1

    #change column names for future convenience
    data_right = ws_right.values
    column_right = ["polygon", "x1", "y1", "x2", "y2", "x3", "y3", "x4", "y4", "price", "lat", "lng"]
    df_right = pd.DataFrame(data_right, columns=column_right)
    df_right = df_right.iloc[1:, 0:12]

    #join the two tables
    pd.set_option('max_columns', None)
    merged_df = pd.merge(df_left, df_right, on="polygon")

    #groupby (mainly) on the polygon_id (other columns listed below for json format)
    #merge all store information of each polygon to a json object
    json_df = (merged_df.groupby(["polygon", "lat", "lng", "x1", "y1", "x2", "y2", "x3", "y3", "x4", "y4"])
                        .apply(lambda x: x[["store_id", "price", "Lat_e7", "Lng_e7",
                                            "Business name", "Address"]].to_dict('records'))
                        .reset_index()
                        .rename(columns={0: "store_info", 1: "poly"})
                        .to_json(orient="records"))

    #this chunk of code mainly aggregates the latitude and longitude of each polygon, as well as changes some of the
    #column names in the json (especially sub-json of the store info) to targeted ones, and append all the store_id's
    #to a specific list in the json object
    count = 0
    my_data = json.loads(json_df)
    for obj in my_data: #json.loads() returns a list, obj is a dict
        storeid_list = []
        position = []
        pos1, pos2, pos3, pos4 = {}, {}, {}, {}
        for item in obj: #obj returns a dict
            if item == "store_info":
                for store in obj[item]: #value returns a list info of stores
                    for item0 in store:
                        if item0 == "store_id":
                            storeid_list.append(store[item0])
                    store["id"] = store["store_id"]
                    del store["store_id"]
                    store["lat"] = store["Lat_e7"]
                    del store["Lat_e7"]
                    store["lng"] = store["Lng_e7"]
                    del store["Lng_e7"]
                    store["biz_name"] = store["Business name"]
                    del store["Business name"]
                    store["addr"] = store["Address"]
                    del store["Address"]
            if (item[0] == 'x' or item[0] == 'y') and item[1] != '_':
                if item == "x1":
                    pos1["lat"] = obj[item]
                if item == "y1":
                    pos1["lng"] = obj[item]
                if item == "x2":
                    pos2["lat"] = obj[item]
                if item == "y2":
                    pos2["lng"] = obj[item]
                if item == "x3":
                    pos3["lat"] = obj[item]
                if item == "y3":
                    pos3["lng"] = obj[item]
                if item == "x4":
                    pos4["lat"] = obj[item]
                if item == "y4":
                    pos4["lng"] = obj[item]
        position.extend([pos1, pos2, pos3, pos4])
        obj["poly"] = position
        obj["store_id"] = storeid_list
        del obj["x1"]; del obj["y1"]; del obj["x2"]; del obj["y2"]
        del obj["x3"]; del obj["y3"]; del obj["x4"]; del obj["y4"]

    #print (test) the json table
    final_data = json.dumps(my_data, indent=2)
    print(final_data)

    #output the polygon id and json objects to excel, iteratively assign the value of each cell
    wb_final = Workbook()
    ws_final = wb_final.active
    c1 = ws_final.cell(row=1, column=1)
    c1.value = "Corpus Code"
    c2 = ws_final.cell(row=1, column=2)
    c2.value = "Script Attributes"
    count_row = 2
    for obj in my_data: #mydata is a list (json array), so obj is a dict (json object)
        corpus_code = changeKeyName(obj["polygon"])
        ws_final.cell(row=count_row, column=1).value = corpus_code
        del obj["polygon"]
        ws_final.cell(row=count_row, column=2).value = json.dumps(obj)
        count_row += 1
    wb_final.save("D:\Jun22\output\myfinal_output.xlsx")
